"""xlsx / xlsm 抽出。

- セル値・シート名: openpyxl（read_only モード）— セル単位で Segment 化
- セルコメント（従来メモ）: xl/comments*.xml をパース（シート関連付け）
- スレッドコメント（Excel 2016+ / Office 2024 既定）:
  xl/threadedComments/threadedComment*.xml + xl/persons/person*.xml をパース
- テキストボックス/図形/SmartArt/グラフ/ワードアート: xl/drawings/* をパース
  → シート rels から関連付けたシート名 + shape 名で locator 化
"""
from __future__ import annotations

import warnings
import zipfile
from typing import Dict, List

from lxml import etree

from search import Segment


# openpyxl が出す「機能未サポート」系の UserWarning を抑制する。
# どれも docgrep の検索対象（セル値・コメント・図形等）には影響しないため
# 実害はないが、走査時に大量に出てログを汚すので message パターンで限定的に抑制。
#
#   - "Cannot parse header or footer so it will be ignored"
#       … 複雑な印刷ヘッダー/フッター書式
#   - "Conditional Formatting extension is not supported and will be removed"
#   - "Data Validation extension is not supported and will be removed"
#   - "Slicer extension is not supported and will be removed"
#   - "Pivot Table extension is not supported and will be removed"
#       … 上記いずれも Office 2010+ の拡張要素 (x14:*) で openpyxl 未対応のもの
warnings.filterwarnings(
    "ignore",
    message=r"Cannot parse header or footer.*",
    category=UserWarning,
)
warnings.filterwarnings(
    "ignore",
    message=r".*extension is not supported.*",
    category=UserWarning,
)


# --- 名前空間 ---
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
A_T = f"{{{A_NS}}}t"
XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
SS_T = f"{{{SS_NS}}}t"
OD_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DRAWING_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
COMMENTS_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments"

# --- スレッドコメント (Excel 2016+ / Office 2024 既定) の名前空間・型 ---
# threadedComment*.xml と persons*.xml は 2018 版 threadedcomments 名前空間、
# sheet ↔ threadedComments の rels は 2017/10 の相対 URI。
TC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
TC_TEXT = f"{{{TC_NS}}}text"
TC_COMMENT = f"{{{TC_NS}}}threadedComment"
TC_PERSON = f"{{{TC_NS}}}person"
THREADED_COMMENTS_REL_TYPE = "http://schemas.microsoft.com/office/2017/10/relationships/threadedComment"


def extract_xlsx(path: str, granularity: str = "cell") -> List[Segment]:
    """xlsx を抽出する。

    granularity:
      - "cell" (既定): 1 セル = 1 Segment。locator は "Sheet1!B5"。
      - "row":  1 行 = 1 Segment（空セル除く各セル値をタブ結合）。
        Segment 数が列数分減り、大規模 xlsx で軽量になる。
        locator は "Sheet1!Row 5"（列番地は失われる）。
    """
    if granularity not in ("cell", "row"):
        granularity = "cell"
    segments: List[Segment] = []

    # 1) セル値・シート名 (openpyxl, read_only でストリーム)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            for ws in wb.worksheets:
                sheet_name = str(getattr(ws, "title", "") or "")
                if sheet_name:
                    segments.append(Segment(
                        text=sheet_name,
                        locator=f"シート名: {sheet_name}",
                    ))
                if granularity == "cell":
                    for row in ws.iter_rows():
                        for cell in row:
                            val = cell.value
                            if val is None:
                                continue
                            segments.append(Segment(
                                text=str(val),
                                locator=f"{sheet_name}!{cell.coordinate}",
                            ))
                else:  # row 粒度
                    row_idx = 0
                    for row in ws.iter_rows(values_only=True):
                        row_idx += 1
                        cells = [str(v) for v in row if v is not None and str(v) != ""]
                        if not cells:
                            continue
                        segments.append(Segment(
                            text="\t".join(cells),
                            locator=f"{sheet_name}!Row {row_idx}",
                        ))
        finally:
            wb.close()
    except Exception:
        # openpyxl で失敗してもオブジェクト抽出は試す
        pass

    # 2) ZIP 内 XML（drawings / charts / diagrams / comments / threadedComments）
    try:
        with zipfile.ZipFile(path) as z:
            name_to_lower = {n: n.lower() for n in z.namelist()}
            lower_to_name = {v: k for k, v in name_to_lower.items()}
            drawing_to_sheet = _build_drawing_sheet_map(z, lower_to_name)
            comments_to_sheet = _build_comments_sheet_map(z, lower_to_name)
            threaded_to_sheet = _build_threaded_comments_sheet_map(z, lower_to_name)
            persons = _load_persons(z, lower_to_name)

            for name, lname in name_to_lower.items():
                if not lname.endswith(".xml"):
                    continue

                if lname.startswith("xl/drawings/drawing"):
                    sheet = drawing_to_sheet.get(lname, "")
                    segments.extend(_extract_drawing(z, name, sheet))

                elif lname.startswith("xl/charts/chart"):
                    chart_id = name.rsplit("/", 1)[-1]
                    text = _gather_drawingml_text(z.read(name))
                    if text:
                        segments.append(Segment(
                            text=text,
                            locator=f"グラフ ({chart_id})",
                        ))

                elif lname.startswith("xl/diagrams/"):
                    diagram_id = name.rsplit("/", 1)[-1]
                    text = _gather_drawingml_text(z.read(name))
                    if text:
                        segments.append(Segment(
                            text=text,
                            locator=f"SmartArt ({diagram_id})",
                        ))

                elif lname.startswith("xl/threadedcomments/threadedcomment"):
                    sheet = threaded_to_sheet.get(lname, "")
                    segments.extend(_extract_threaded_comments(
                        z.read(name), sheet, persons,
                    ))

                elif lname.startswith("xl/comments"):
                    sheet = comments_to_sheet.get(lname, "")
                    segments.extend(_extract_comments(z, name, sheet))
    except (zipfile.BadZipFile, OSError):
        pass

    return segments


# -----------------------------------------------------------------------------
# シート関連付け（drawings / comments）
# -----------------------------------------------------------------------------
def _build_sheet_name_to_part(z: zipfile.ZipFile) -> Dict[str, str]:
    """sheet 表示名 → "xl/worksheets/sheet1.xml" (小文字) のマップ。"""
    try:
        wb_xml = etree.fromstring(z.read("xl/workbook.xml"))
        rels_xml = etree.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    except (KeyError, etree.XMLSyntaxError):
        return {}

    rid_to_target: Dict[str, str] = {}
    for rel in rels_xml.findall(f"{{{PKG_REL_NS}}}Relationship"):
        rid_to_target[rel.get("Id", "")] = rel.get("Target", "")

    result: Dict[str, str] = {}
    for sheet in wb_xml.findall(f"{{{SS_NS}}}sheets/{{{SS_NS}}}sheet"):
        name = sheet.get("name", "")
        rid = sheet.get(f"{{{OD_REL_NS}}}id", "")
        target = rid_to_target.get(rid, "")
        if not name or not target:
            continue
        if target.startswith("/"):
            full = target.lstrip("/")
        else:
            full = "xl/" + target
        result[name] = full.lower()
    return result


def _sheet_rels_path(sheet_part: str) -> str:
    """xl/worksheets/sheet1.xml → xl/worksheets/_rels/sheet1.xml.rels"""
    parts = sheet_part.split("/")
    parts.insert(-1, "_rels")
    parts[-1] = parts[-1] + ".rels"
    return "/".join(parts)


def _resolve_target(base: str, target: str) -> str:
    """rels の相対 Target を ZIP 内パス（小文字）に解決する。

    base: xl/worksheets/sheet1.xml / target: ../drawings/drawing1.xml
    → xl/drawings/drawing1.xml
    """
    if target.startswith("/"):
        return target.lstrip("/").lower()
    base_dir = base.rsplit("/", 1)[0]
    parts = base_dir.split("/")
    for tok in target.split("/"):
        if tok == "..":
            if parts:
                parts.pop()
        elif tok and tok != ".":
            parts.append(tok)
    return "/".join(parts).lower()


def _build_drawing_sheet_map(z: zipfile.ZipFile, lower_to_name: Dict[str, str]) -> Dict[str, str]:
    sheet_to_part = _build_sheet_name_to_part(z)
    out: Dict[str, str] = {}
    for sheet_name, part in sheet_to_part.items():
        rels_lower = _sheet_rels_path(part)
        orig = lower_to_name.get(rels_lower)
        if not orig:
            continue
        try:
            rels_xml = etree.fromstring(z.read(orig))
        except (KeyError, etree.XMLSyntaxError):
            continue
        for rel in rels_xml.findall(f"{{{PKG_REL_NS}}}Relationship"):
            if rel.get("Type") == DRAWING_REL_TYPE:
                resolved = _resolve_target(part, rel.get("Target", ""))
                out[resolved] = sheet_name
    return out


def _build_comments_sheet_map(z: zipfile.ZipFile, lower_to_name: Dict[str, str]) -> Dict[str, str]:
    sheet_to_part = _build_sheet_name_to_part(z)
    out: Dict[str, str] = {}
    for sheet_name, part in sheet_to_part.items():
        rels_lower = _sheet_rels_path(part)
        orig = lower_to_name.get(rels_lower)
        if not orig:
            continue
        try:
            rels_xml = etree.fromstring(z.read(orig))
        except (KeyError, etree.XMLSyntaxError):
            continue
        for rel in rels_xml.findall(f"{{{PKG_REL_NS}}}Relationship"):
            if rel.get("Type") == COMMENTS_REL_TYPE:
                resolved = _resolve_target(part, rel.get("Target", ""))
                out[resolved] = sheet_name
    return out


def _build_threaded_comments_sheet_map(z: zipfile.ZipFile, lower_to_name: Dict[str, str]) -> Dict[str, str]:
    """threadedComment*.xml (小文字パス) → sheet 表示名 のマップ。"""
    sheet_to_part = _build_sheet_name_to_part(z)
    out: Dict[str, str] = {}
    for sheet_name, part in sheet_to_part.items():
        rels_lower = _sheet_rels_path(part)
        orig = lower_to_name.get(rels_lower)
        if not orig:
            continue
        try:
            rels_xml = etree.fromstring(z.read(orig))
        except (KeyError, etree.XMLSyntaxError):
            continue
        for rel in rels_xml.findall(f"{{{PKG_REL_NS}}}Relationship"):
            if rel.get("Type") == THREADED_COMMENTS_REL_TYPE:
                resolved = _resolve_target(part, rel.get("Target", ""))
                out[resolved] = sheet_name
    return out


def _load_persons(z: zipfile.ZipFile, lower_to_name: Dict[str, str]) -> Dict[str, str]:
    """xl/persons/person*.xml をパースして personId → displayName マップを返す。

    threadedComment の personId 属性はここに定義された `id` と対応する。
    """
    persons: Dict[str, str] = {}
    for lname, orig in lower_to_name.items():
        if not lname.startswith("xl/persons/") or not lname.endswith(".xml"):
            continue
        try:
            root = etree.fromstring(z.read(orig))
        except (KeyError, etree.XMLSyntaxError):
            continue
        for p in root.iter(TC_PERSON):
            pid = p.get("id", "")
            name = p.get("displayName", "")
            if pid:
                persons[pid] = name
    return persons


# -----------------------------------------------------------------------------
# drawings / comments の XML パース
# -----------------------------------------------------------------------------
def _extract_drawing(z: zipfile.ZipFile, name: str, sheet: str) -> List[Segment]:
    segs: List[Segment] = []
    try:
        root = etree.fromstring(z.read(name))
    except etree.XMLSyntaxError:
        return segs

    sheet_prefix = (sheet + ": ") if sheet else ""
    drawing_id = name.rsplit("/", 1)[-1]

    # xdr:sp = shape (テキストボックス・図形)
    for sp in root.iter(f"{{{XDR_NS}}}sp"):
        shape_name = _get_shape_name(sp, "nvSpPr")
        text = _gather_drawingml_text_from_elem(sp)
        if not text:
            continue
        if shape_name:
            loc = f"{sheet_prefix}図形「{shape_name}」"
        else:
            loc = f"{sheet_prefix}図形 ({drawing_id})"
        segs.append(Segment(text=text, locator=loc))

    # xdr:graphicFrame = chart / SmartArt / その他埋め込み
    for gf in root.iter(f"{{{XDR_NS}}}graphicFrame"):
        gname = _get_shape_name(gf, "nvGraphicFramePr")
        text = _gather_drawingml_text_from_elem(gf)
        if not text:
            continue
        if gname:
            loc = f"{sheet_prefix}グラフ/図解「{gname}」"
        else:
            loc = f"{sheet_prefix}グラフ/図解 ({drawing_id})"
        segs.append(Segment(text=text, locator=loc))

    return segs


def _get_shape_name(elem, wrapper_local: str) -> str:
    cnvpr = elem.find(f"{{{XDR_NS}}}{wrapper_local}/{{{XDR_NS}}}cNvPr")
    if cnvpr is not None:
        return cnvpr.get("name", "")
    return ""


def _gather_drawingml_text_from_elem(elem) -> str:
    parts = []
    for t in elem.iter(A_T):
        if t.text:
            parts.append(t.text)
    return "\n".join(parts)


def _gather_drawingml_text(xml_bytes: bytes) -> str:
    try:
        root = etree.fromstring(xml_bytes)
    except etree.XMLSyntaxError:
        return ""
    return _gather_drawingml_text_from_elem(root)


def _extract_comments(z: zipfile.ZipFile, name: str, sheet: str) -> List[Segment]:
    segs: List[Segment] = []
    try:
        root = etree.fromstring(z.read(name))
    except etree.XMLSyntaxError:
        return segs

    sheet_prefix = (sheet + "!") if sheet else ""
    authors: List[str] = []
    authors_node = root.find(f"{{{SS_NS}}}authors")
    if authors_node is not None:
        for a in authors_node.findall(f"{{{SS_NS}}}author"):
            authors.append(a.text or "")

    list_node = root.find(f"{{{SS_NS}}}commentList")
    if list_node is None:
        return segs
    for c in list_node.findall(f"{{{SS_NS}}}comment"):
        ref = c.get("ref", "")
        author_id = c.get("authorId", "")
        try:
            author = authors[int(author_id)] if author_id else ""
        except (ValueError, IndexError):
            author = ""
        text_parts = []
        for t in c.iter(SS_T):
            if t.text:
                text_parts.append(t.text)
        text = "".join(text_parts).strip()
        if not text:
            continue
        loc = f"{sheet_prefix}{ref} コメント" if ref else "コメント"
        if author:
            loc += f" ({author})"
        segs.append(Segment(text=text, locator=loc))
    return segs


def _extract_threaded_comments(xml_bytes: bytes, sheet: str,
                               persons: Dict[str, str]) -> List[Segment]:
    """threadedComment*.xml をパースして Segment リストを返す。

    parentId 属性の有無で「スレッドコメント」と「スレッド返信」を区別し、
    persons マップから発言者の displayName を解決して locator に埋める。
    """
    segs: List[Segment] = []
    try:
        root = etree.fromstring(xml_bytes)
    except etree.XMLSyntaxError:
        return segs

    sheet_prefix = (sheet + "!") if sheet else ""
    for c in root.iter(TC_COMMENT):
        ref = c.get("ref", "")
        person_id = c.get("personId", "")
        parent_id = c.get("parentId", "")
        author = persons.get(person_id, "")
        text_node = c.find(TC_TEXT)
        if text_node is None or not text_node.text:
            continue
        text = text_node.text.strip()
        if not text:
            continue
        kind = "スレッド返信" if parent_id else "スレッドコメント"
        loc = f"{sheet_prefix}{ref} {kind}" if ref else kind
        if author:
            loc += f" ({author})"
        segs.append(Segment(text=text, locator=loc))
    return segs
