"""xlsx 抽出器の locator 機能テスト（openpyxl で実 xlsx を生成して検証）。"""
from openpyxl import Workbook
from openpyxl.comments import Comment

from extractors.xlsx import extract_xlsx


def _build_xlsx(path, sheets):
    """sheets = {sheet_name: [[cell, ...], ...]} で xlsx を生成する。"""
    wb = Workbook()
    # デフォルトの "Sheet" を削除
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)


# === 基本: セル粒度 ===

def test_cell_granularity_emits_sheet_name_and_cells(tmp_path):
    f = tmp_path / "basic.xlsx"
    _build_xlsx(f, {"Data": [["A1", "B1", "C1"], ["A2", None, "C2"]]})
    segs = extract_xlsx(str(f), granularity="cell")
    locs = [s.locator for s in segs]
    texts = [s.text for s in segs]

    # シート名 segment
    assert "シート名: Data" in locs
    # セル座標 segment（None セルは含まない）
    assert "Data!A1" in locs
    assert "Data!B1" in locs
    assert "Data!C1" in locs
    assert "Data!A2" in locs
    assert "Data!B2" not in locs   # None スキップ
    assert "Data!C2" in locs
    # text 内容も対応
    pairs = dict(zip(locs, texts))
    assert pairs["Data!B1"] == "B1"


def test_multiple_sheets_use_each_sheet_name(tmp_path):
    f = tmp_path / "multi.xlsx"
    _build_xlsx(f, {"売上": [["合計", 100]], "費用": [["合計", 50]]})
    segs = extract_xlsx(str(f), granularity="cell")
    locs = [s.locator for s in segs]
    assert "シート名: 売上" in locs
    assert "シート名: 費用" in locs
    assert "売上!A1" in locs
    assert "費用!B1" in locs


def test_japanese_sheet_name_and_value(tmp_path):
    f = tmp_path / "ja.xlsx"
    _build_xlsx(f, {"東京支社": [["プロジェクトA", "売上目標"]]})
    segs = extract_xlsx(str(f), granularity="cell")
    locs = [s.locator for s in segs]
    texts = [s.text for s in segs]
    assert "シート名: 東京支社" in locs
    assert "東京支社!A1" in locs
    assert "プロジェクトA" in texts


# === 行粒度 ===

def test_row_granularity_aggregates_row_values(tmp_path):
    f = tmp_path / "row.xlsx"
    _build_xlsx(f, {"R": [["a", "b", "c"], ["d", None, "f"], [None, None, None]]})
    segs = extract_xlsx(str(f), granularity="row")
    # 行 segment は空行を除く 2 件 + シート名 1 件
    row_segs = [s for s in segs if s.locator.startswith("R!Row ")]
    locs = [s.locator for s in row_segs]
    texts = [s.text for s in row_segs]
    assert "R!Row 1" in locs
    assert "R!Row 2" in locs
    # 3 行目は全 None なのでスキップ
    assert "R!Row 3" not in locs
    # タブ区切りで結合される
    r1 = dict(zip(locs, texts))["R!Row 1"]
    assert r1 == "a\tb\tc"


def test_unknown_granularity_falls_back_to_cell(tmp_path):
    f = tmp_path / "fb.xlsx"
    _build_xlsx(f, {"X": [["one"]]})
    segs = extract_xlsx(str(f), granularity="badvalue")
    locs = [s.locator for s in segs]
    # cell 粒度として処理される
    assert "X!A1" in locs


# === コメント ===

def test_cell_comment_extracted_with_locator(tmp_path):
    f = tmp_path / "cm.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Notes")
    ws["B5"] = "本体"
    ws["B5"].comment = Comment("これは検討事項のメモです", "Yamada")
    wb.save(f)
    segs = extract_xlsx(str(f), granularity="cell")
    locs = [s.locator for s in segs]
    texts = [s.text for s in segs]
    assert "Notes!B5" in locs
    # コメント由来の segment が含まれる（locator に "コメント" を含むはず）
    has_comment_seg = any("コメント" in l and "B5" in l for l in locs)
    assert has_comment_seg, f"コメント segment が見つかりません。locs={locs}"
    has_comment_text = any("検討事項" in t for t in texts)
    assert has_comment_text


def test_comment_author_in_locator(tmp_path):
    f = tmp_path / "cm_author.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="A")
    ws["A1"] = "セル"
    ws["A1"].comment = Comment("注記", "Tanaka")
    wb.save(f)
    segs = extract_xlsx(str(f), granularity="cell")
    # author が locator に含まれる
    assert any("Tanaka" in s.locator for s in segs), \
        f"author 'Tanaka' が locator に含まれていません。locs={[s.locator for s in segs]}"


# === 図形・テキストボックス ===
# openpyxl の Image/Drawing 操作は read-only に近く、Drawing 配置 + テキスト
# 注入は煩雑なため、ここではセル・コメントの locator にフォーカスする。
# drawings のテストは将来、別途実 xlsx fixture を導入して検証する。


# === 破損ファイル耐性 ===

def test_corrupted_xlsx_returns_empty_without_raising(tmp_path):
    f = tmp_path / "broken.xlsx"
    f.write_bytes(b"not a real xlsx")
    # zip でも openpyxl でもない → 例外を握りつぶして空リスト
    segs = extract_xlsx(str(f), granularity="cell")
    assert segs == []


# =============================================================================
# スレッドコメント (Excel 2016+ / Office 2024 既定)
# =============================================================================

_THREADED_XML_STANDARD = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<ThreadedComments xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">
    <threadedComment ref="B5" personId="{PID1}" id="{ID1}">
        <text>これは検討中の項目です</text>
    </threadedComment>
    <threadedComment ref="B5" personId="{PID2}" id="{ID2}" parentId="{ID1}">
        <text>了解しました</text>
    </threadedComment>
    <threadedComment ref="D8" personId="{PID3}" id="{ID3}">
        <text>単発コメント（発言者マップに無い personId）</text>
    </threadedComment>
</ThreadedComments>""".encode("utf-8")


def test_threaded_comments_parse_with_persons():
    from extractors.xlsx import _extract_threaded_comments
    persons = {"{PID1}": "山田太郎", "{PID2}": "鈴木花子"}
    segs = _extract_threaded_comments(_THREADED_XML_STANDARD, "Sheet1", persons)
    assert len(segs) == 3

    # 最初のコメント（parentId 無し）→ スレッドコメント
    assert segs[0].text == "これは検討中の項目です"
    assert segs[0].locator == "Sheet1!B5 スレッドコメント (山田太郎)"

    # parentId あり → スレッド返信
    assert segs[1].text == "了解しました"
    assert segs[1].locator == "Sheet1!B5 スレッド返信 (鈴木花子)"

    # persons に無い personId → 発言者名なし
    assert segs[2].text == "単発コメント（発言者マップに無い personId）"
    assert segs[2].locator == "Sheet1!D8 スレッドコメント"


def test_threaded_comments_without_sheet_name():
    from extractors.xlsx import _extract_threaded_comments
    segs = _extract_threaded_comments(_THREADED_XML_STANDARD, "", {})
    # sheet 空 → prefix 無しでも動作
    assert segs[0].locator == "B5 スレッドコメント"


def test_threaded_comments_empty_text_is_skipped():
    from extractors.xlsx import _extract_threaded_comments
    xml = ("""<?xml version="1.0"?>
<ThreadedComments xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">
    <threadedComment ref="A1" personId="{X}" id="{Y}"><text></text></threadedComment>
    <threadedComment ref="A2" personId="{X}" id="{Z}"><text>   </text></threadedComment>
    <threadedComment ref="A3" personId="{X}" id="{W}"><text>OK</text></threadedComment>
</ThreadedComments>""").encode("utf-8")
    segs = _extract_threaded_comments(xml, "S1", {})
    # 空文字 / 空白のみは除外、"OK" だけ残る
    assert len(segs) == 1
    assert segs[0].text == "OK"


def test_threaded_comments_invalid_xml_returns_empty():
    from extractors.xlsx import _extract_threaded_comments
    assert _extract_threaded_comments(b"not xml", "Sheet1", {}) == []


def test_load_persons_from_zip(tmp_path):
    """xl/persons/person*.xml が読み込まれて id→displayName になること。"""
    import zipfile as _zf
    from extractors.xlsx import _load_persons

    xlsx_path = tmp_path / "with_persons.xlsx"
    persons_xml = ("""<?xml version="1.0"?>
<personList xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">
    <person id="{P1}" displayName="山田太郎" userId="s::y@example.com" providerId="AD"/>
    <person id="{P2}" displayName="鈴木花子" userId="s::h@example.com" providerId="AD"/>
</personList>""").encode("utf-8")
    with _zf.ZipFile(xlsx_path, "w") as z:
        z.writestr("xl/persons/person1.xml", persons_xml)

    with _zf.ZipFile(xlsx_path) as z:
        name_to_lower = {n: n.lower() for n in z.namelist()}
        lower_to_name = {v: k for k, v in name_to_lower.items()}
        persons = _load_persons(z, lower_to_name)

    assert persons == {"{P1}": "山田太郎", "{P2}": "鈴木花子"}
