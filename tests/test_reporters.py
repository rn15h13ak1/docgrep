"""reporter の round-trip 出力テスト。

Excel は openpyxl で読み返してシート構造を確認、HTML は文字列として
キーワード・モーダル・エラー一覧の有無を確認する。
"""
from search import FileResult, Hit, Searcher


def _make_hits():
    fr = FileResult(
        path="C:/share/foo.md",
        hits=[
            Hit(snippet="...東京本社では...", position=10, matched="東京",
                locator="行 12"),
            Hit(snippet="...プロジェクトA...", position=40, matched="プロジェクト",
                locator="行 15"),
        ],
    )
    err = FileResult(
        path="C:/share/broken.docx",
        error="extract_error: corrupted",
    )
    return [fr], [err]


def _summary():
    return {
        "走査ファイル数": 5,
        "ヒットファイル数": 1,
        "ヒット件数合計": 2,
        "スキップ数": 1,
        "エラー数": 1,
    }


def test_excel_writes_results_summary_errors_skipped(tmp_path):
    from openpyxl import load_workbook
    from reporter.excel import write_excel

    hits, errs = _make_hits()
    out = tmp_path / "r.xlsx"
    skipped = {"binary": ["C:/share/a.png", "C:/share/b.jpg"]}
    write_excel(str(out), hits, _summary(), errors=errs, skipped=skipped)

    wb = load_workbook(str(out))
    assert {"results", "summary", "errors", "skipped"}.issubset(set(wb.sheetnames))
    ws = wb["results"]
    # ヘッダ + ヒット 2 行 = 計 3 行
    assert ws.max_row == 3
    # ヘッダ列名のいずれかが期待のもの
    headers = [c.value for c in ws[1]]
    assert "検知箇所" in headers and "ヒット語" in headers

    ws_skipped = wb["skipped"]
    rows = [tuple(r) for r in ws_skipped.iter_rows(values_only=True)]
    assert rows[0] == ("理由", "パス")
    assert ("binary", "C:/share/a.png") in rows


def test_excel_sanitises_control_chars(tmp_path):
    from openpyxl import load_workbook
    from reporter.excel import write_excel

    fr = FileResult(
        path="C:/share/foo.md",
        hits=[
            Hit(
                snippet="bel\x07vt\x0bend",
                position=0,
                matched="bel",
                locator="行 1",
            ),
        ],
    )
    out = tmp_path / "r.xlsx"
    write_excel(str(out), [fr], _summary())
    wb = load_workbook(str(out))
    ws = wb["results"]
    snippet_col = [c.value for c in ws["E"]][1]  # 1行目はヘッダ
    assert "\x07" not in snippet_col
    assert "\x0b" not in snippet_col


def test_html_contains_hits_and_modal(tmp_path):
    from reporter.html import write_html

    searcher = Searcher(
        mode="keyword",
        keywords=["東京", "プロジェクト"],
        operator="or",
        case_sensitive=False,
        normalize_width=True,
        fuzzy_threshold=0.8,
        snippet_chars=30,
        max_hits_per_file=10,
    )
    hits, errs = _make_hits()
    skipped = {"binary": ["C:/share/a.png"], "scope_out": ["C:/share/b.pdf"]}
    out = tmp_path / "r.html"
    write_html(str(out), hits, _summary(), searcher, errors=errs, skipped=skipped)

    text = out.read_text(encoding="utf-8")
    assert "docgrep 検索結果" in text
    assert "東京" in text  # ハイライト対象
    assert 'class="hl"' in text
    assert 'class="loc"' in text
    # モーダル関連
    assert 'id="skip-dialog"' in text
    assert 'data-reason="binary"' in text
    assert 'data-reason="scope_out"' in text
    # エラー一覧
    assert "エラー一覧" in text and "broken.docx" in text


def test_html_omits_modal_when_no_skipped(tmp_path):
    from reporter.html import write_html

    searcher = Searcher(
        mode="keyword", keywords=["x"], operator="or",
        case_sensitive=False, normalize_width=True,
        fuzzy_threshold=0.8, snippet_chars=10, max_hits_per_file=10,
    )
    hits, _ = _make_hits()
    out = tmp_path / "r2.html"
    write_html(str(out), hits, _summary(), searcher)
    text = out.read_text(encoding="utf-8")
    # skipped 未指定なのでモーダル要素は含まれない
    assert 'id="skip-dialog"' not in text
