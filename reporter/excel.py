"""Excel 出力（openpyxl）。

results シートは 1 ヒット 1 行（パスの繰り返しは許容）で、検知箇所
(locator) を独立カラムに持たせる。オートフィルタ・並べ替えがしやすい。

注意: openpyxl は OOXML 仕様の制限から一部の制御文字 (\x00-\x08, \x0B, \x0C,
\x0E-\x1F) を含むセル値を受け付けず IllegalCharacterError で wb.save() が
失敗する。Word/PowerPoint から COM で抜いた本文に \x07 (BEL) や \x0B (VT)
等が紛れることがあるため、append 前に必ず _sanitize() を通す。
"""
from __future__ import annotations

import os
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

from search import FileResult


def _sanitize(val: Any) -> Any:
    """openpyxl が拒否する制御文字を半角スペースに置換する。

    文字列以外（数値・None など）はそのまま返す。
    """
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub(" ", val)
    return val


def _row(*cells: Any) -> List[Any]:
    return [_sanitize(c) for c in cells]


def write_excel(
    out_path: str,
    file_results: Iterable[FileResult],
    summary: Dict[str, object],
    errors: Optional[Iterable[FileResult]] = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    headers = ["パス", "拡張子", "検知箇所", "ヒット語", "スニペット", "最終更新日時"]
    ws.append(headers)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for fr in file_results:
        if not fr.hits:
            continue
        ext = os.path.splitext(fr.path)[1].lower()
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(fr.path)).strftime("%Y-%m-%d %H:%M:%S")
        except OSError:
            mtime = ""
        for hit in fr.hits:
            ws.append(_row(fr.path, ext, hit.locator, hit.matched, hit.snippet, mtime))
            ws.cell(row=ws.max_row, column=5).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # サマリーシート
    ws2 = wb.create_sheet("summary")
    ws2.append(["項目", "値"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for k, v in summary.items():
        ws2.append(_row(k, v))
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 60

    # エラーシート（抽出 / 検索失敗ファイルの内訳）
    err_list = [fr for fr in (errors or []) if fr.error]
    if err_list:
        ws3 = wb.create_sheet("errors")
        ws3.append(["パス", "エラー内容"])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for fr in err_list:
            ws3.append(_row(fr.path, fr.error))
        ws3.column_dimensions["A"].width = 80
        ws3.column_dimensions["B"].width = 80
        ws3.freeze_panes = "A2"

    wb.save(out_path)
